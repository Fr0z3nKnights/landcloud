"""
@author:Fr0z3n
@contact:websec@yeah.net
@datetime:2020/3/16
@desc:To get json data from url(https://landcloud.org.cn),and distribute personal task.
"""

from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import os
import datetime
import re
# 屏蔽SSL报错,要求重新安装requests==2.7.0(version)
from requests.packages import urllib3
urllib3.disable_warnings()


class Tysdgx(object):
    def __init__(self, pagecnt=52, assignfile="./tysdgx_data/assign.txt"):
        # 总页数52页，每页100个编号
        self.pagecnt = pagecnt
        self.vlist_url = "https://jg.landcloud.org.cn:8090/webapi/api/vlist"
        self.assign_url = "https://jg.landcloud.org.cn:8090/webapi/api//TaskDistribute"
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:74.0) Gecko/20100101 Firefox/74.0",
            "Authorization": "",# token cannot be null
            "Origin": "https://jg.landcloud.org.cn:8090",
            "Referer": "https://jg.landcloud.org.cn:8090/main/list/tysdgx/1/0",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
            "Accept-Encoding": "gzip, deflate, br",
            "Content-Type": "application/json;charset=utf-8",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
        }
        self.cookies = {"482b4337498e47ae9e76f0e2271caf89": "WyIzODA1NTY1NTEwIl0"}
        self.data = {
            "ywlx": "tysdgx",
            "xzqdm": "410922",
            "level": 1,
            "userxzqdm": "410922",
            "pageSize": 100,
            "pageIndex": 1,
            "order": "",
            "isUseCacheCount": True
        }
        self.tbbsm_li = {
            "tbbsm": "",
            "xmbh": "",
            "xzqdm": "410922"
        }
        self.taskdata ={
            "tbbsms": [],
            "xzqdm": "410922",
            "ywlx": "TYSDGX",
            "sbbsms": [614864, 615084, 526084, 550264, 555304, 555284, 555264, 555504, 555604, 556624, 555324, 556264]
        }
        self.userdata = {
            "username": "",
            "password": "",
            "verifyCode": ""
        }
        self.maxjctb = 5161
        self.update_flag = True
        self.assign_name = assignfile
        self.datapath = "./tysdgx_data/TYSDGX_ALL_JCTB.xlsx"
        self.auth = ''

    # 登录
    def landlogin(self):
        if not os.path.exists("./tysdgx_data"):
            os.mkdir("./tysdgx_data")
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:74.0) Gecko/20100101 Firefox/74.0"
        }

        host_url = 'https://landcloud.org.cn'
        se = requests.session()
        # 获取索引页面设置cookies
        index = se.get(url=host_url + "/index", headers=headers, verify=False)
        print("[ %s ]++网络状态：" % self.hms_now(), index.status_code)
        if index.status_code == 200:
            # , index.url, index.headers)
            # print(response.content)
            # dic = self.cookiejar2dic(se.cookies)
            # print(dic)
            # 获取登录页面
            getlogin = se.get(url=host_url + "/login", cookies=se.cookies, verify=False)
            print("[ %s ]++获取登录页面状态：getlogin status:" % self.hms_now(), getlogin.status_code)
            patten = "<img id=\"verifycodeImage\"\s+src=\"(.*?)\"/>"
            # 匹配登录页面上验证码图片的链接
            match = re.findall(pattern=patten, string=getlogin.text)
            print("[ %s ]++匹配到验证码链接：" % self.hms_now(), match[0])
            veryimg_url = host_url + match[0]
            dic2 = self.cookiejar2dic(se.cookies)
            # print(veryimg_url)
            print("[ %s ]++登录页面获得cookies：" % self.hms_now(), dic2)

            # 获取验证码图片
            get_verify_img = se.get(url=veryimg_url, cookies=se.cookies, headers=headers, verify=False)
            print("[ %s ]++获取验证码链接图片：" % self.hms_now(), get_verify_img.status_code)
            # print(get_verify_img.text)
            dic3 = self.cookiejar2dic(se.cookies)
            vfpath = "./tysdgx_data/%s.png" % (datetime.datetime.now().strftime("%y%m%d-%H%M%S") + "_" +dic3['timestamp'])
            # 将获取图片写入文件
            with open(vfpath, "wb") as fp:
                fp.write(get_verify_img.content)
            print("[ %s ]++更新页面获得cookies：" % self.hms_now(), dic3)

            # 发送登录消息，等待登录验证通过
            verifycode = input("[ %s ]++Please input the code you have identified:" % self.hms_now())
            verifycode = verifycode.strip()
            self.userdata['verifyCode'] = verifycode
            if os.path.exists(vfpath):  # 删除生成的.png文件
                os.remove(vfpath)
                print("[ %s ]++文件路径：%s" % (self.hms_now(), vfpath), "删除成功！")

            signin_url = host_url + "/login/login.action"
            getsignin = se.post(url=signin_url, cookies=se.cookies, params=self.userdata, headers=headers, verify=False)
            # print(getsignin.url)
            if getsignin.status_code == 200 and getsignin.json()['status'] == 'OK':
                print("[ %s ]++登录页面返回信息：" % self.hms_now(), getsignin.json())

            # 发送跨域消息，获取跨域的token即header中的Authorization
            auth_url = host_url + "/third/proxy/getListDetailPageUrl?ywlx=TYSDGX&type=1&xzqdm=410922"
            get_auth = se.get(url=auth_url, cookies=se.cookies, headers=headers, verify=False)
            if get_auth.status_code == 200 and get_auth.json()['status'] == 'OK':
                self.auth = re.findall(r"token=(.*?)$", get_auth.json()['data'])
                print("[ %s ]++获取跨域token：" % self.hms_now(), self.auth)
            return self.auth[0]
        else:
            return False

    # cookiejar object to dictionary
    def cookiejar2dic(self, cookies):
        return requests.utils.dict_from_cookiejar(cookies)

    # retrieve data from server
    def rtv_data(self):
        # 调用登录流程
        self.headers['Authorization'] = "bearer " + self.landlogin()
        print("[ %s ]++开始获取来自网络的数据并更新：" % self.hms_now())
        filepath = self.datapath
        jcbh2tbbsm = {}
        # 判断是否存在文件
        if os.path.exists(filepath):
            wb = load_workbook(filepath)
            ws = wb.active
            # 判断是否需要更新文件
            if self.maxjctb+1 <= ws.max_row:
                print("[ %s ]++Excel中最大行数为%s,不需要更新数据！" % (self.hms_now(), ws.max_row))
                wb.close()
                return
            else:
                self.update_flag = True
                a = 1
                for row in ws.iter_rows():
                    if a == 1:
                        a += 1
                        continue
                    else:
                        # excel中workbook加载以后，下标从0开始，row[i].value为值
                        # print("+第%s行" % a, row[1].value, row[2].value)
                        jcbh2tbbsm[row[2].value] = row[1].value
                        a += 1
                # print(jcbh2tbbsm)
                print("[ %s ]网络数据大于本地数据,开始更新数据到本地：" % self.hms_now())
        else:
            self.update_flag = False
            print("[ %s ]本地不存在excel数据,开始创建数据并更新,请注意检查数据页数！" % self.hms_now())
            wb = Workbook()
            # 激活 worksheet
            ws = wb.active
            # worksheet rename
            ws.title = "TYSDGX_all_jctb"  # + datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            # insert into worksheet by one row.
            # Excel中的行列顺序
            ws.append(["bsm", "tbbsm", "jcbh", "xzb", "yzb", "tblx", "wyzt"])

        i = 1
        for j in range(1, self.pagecnt+1):
            self.data['pageIndex'] = j
            r = requests.post(url=self.vlist_url, headers=self.headers,
                              json=self.data, cookies=self.cookies, verify=False)
            # print(r.cookies)
            # check if the back status is 200, and retrieve the json data.
            if r.status_code == 200:
                jdata = r.json()
                # count = jdata['data']['allcount']  # int,the count number of all the records.
                record = jdata['data']['records']  # list,and each one in list was a dict.
                for li in record:
                    tbbsm = li['tbbsm']  # 图斑标识码
                    jcbh = li['jcbh']  # 监察编号
                    tblx = li['tblx']
                    bsm = li['bsm']
                    xzb = li['xzb']
                    yzb = li['yzb']
                    wyzt = li['wyzt']
                    if self.update_flag:
                        if jcbh in jcbh2tbbsm.keys():
                            i += 1
                            continue
                        else:
                            ws.append([bsm, tbbsm, jcbh, xzb, yzb, tblx, wyzt])
                            print("[ %s ]+添加第%s行" % (self.hms_now(), i), bsm, tbbsm, jcbh, xzb, yzb, tblx, wyzt)
                            i += 1
                    else:
                        ws.append([bsm, tbbsm, jcbh, xzb, yzb, tblx, wyzt])
                        print("[ %s ]+第%s行" % (self.hms_now(), i), bsm, tbbsm, jcbh, xzb, yzb, tblx, wyzt)
                        i += 1
            else:
                print("[ %s ]+请求数据状态码：%s,退出！" % (self.hms_now(), r.status_code))
        wb.save(filepath)
        wb.close()
        print("[ %s ]+文件保存成功！位于%s." % (self.hms_now(), filepath))
    # get nowtime in H:M:S

    def hms_now(self):
        return datetime.datetime.now().strftime("%H:%M:%S")
    # get jctb from .txt

    def data_matching(self):
        print("[ %s ]++开始处理分发任务：" % self.hms_now())
        print("[ %s ]++检查任务分发文件是否存在重复图斑?" % self.hms_now())
        assignfilepath = self.assign_name
        all_tb = {}
        toassign = {}  # 要分配任务的图斑字典
        notfound = []  # 没发现的图斑列表
        # 检查分配任务图斑文件是否存在
        if os.path.exists(assignfilepath):
            with open(assignfilepath, "r") as fp:
                assign_tb = fp.readlines()
                # print(assign_tb)
                for tb in assign_tb:
                    tb = tb.strip()
        else:
            print("[ %s ]++不存在相关分配任务文件！" % self.hms_now())
            return False
        # 检查任务分配文件图斑重复项并处理
        if len(set(assign_tb)) == len(assign_tb):
            print("[ %s ]++检查结果：任务分发文件图斑无重复项." % self.hms_now())
        else:
            print("[ %s ]**********检查结果：分发任务文件图斑存在重复项,请检查！" % self.hms_now())
            assign_tb = set(assign_tb)
            print("[ %s ]++经处理后图斑列表为：！" % (self.hms_now()), assign_tb)
        # 检查本地数据库文件是否存在
        if os.path.exists(self.datapath):
            wb = load_workbook(self.datapath)
            ws = wb.active
            b = 1
            for row in ws.iter_rows():
                if b == 1:
                    b += 1
                    continue
                else:
                    all_tb[row[2].value] = row[1].value
                    b += 1
            wb.close()
            print("[ %s ]++读取本地数据库长度：%s." % (self.hms_now(), len(all_tb)))

            # 开始匹配从txt读取图斑数是否都存在于网络端
            for tb in assign_tb:
                tb = tb.strip()
                if tb in all_tb.keys():
                    toassign[tb] = all_tb[tb]
                else:
                    print("[ %s ]+图斑%s未找到！" % (self.hms_now(), tb))
                    notfound.append(tb)
            # 对于匹配结果进行处理
            if len(assign_tb) == len(toassign):
                # print("[ %s ]++匹配到图斑字典长度：" % self.hms_now(, toassign)
                print("[ %s ]++匹配到全部要分发图斑共%s个,开始分发任务:" % (self.hms_now(), len(toassign)))
            else:
                print("[ %s ]**********未全部匹配到图斑数！" % self.hms_now())
                nffilepath = "./tysdgx_data/notfound" + datetime.datetime.now().strftime("%y%m%d_%H%M%S") + ".txt"
                with open(nffilepath, "w") as fp:
                        fp.writelines(notfound)
                print("[ %s ]**********未匹配图斑文件位于：%s！" % (self.hms_now(), nffilepath))
                return False
        else:
            print("[ %s ]**********不存在本地数据库文件%s,请检查！" % (self.hms_now(), self.datapath))
            return False
        # 要分配任务图斑在toassign,key为jcbh,value为tbbsm
        return toassign
    # 开始分发任务图斑到移动端

    def begin_assignment(self):
        # 先匹配数据
        assigndata = self.data_matching()
        c = 1
        flag = 1
        if assigndata:
            remain = divmod(len(assigndata), 50)
            for tb in assigndata.keys():
                self.tbbsm_li['xmbh'] = tb
                self.tbbsm_li['tbbsm'] = assigndata[tb]
                self.taskdata['tbbsms'].append(self.tbbsm_li)
                if (remain[0] == 0 and c == remain[1]) or (remain[0] < flag and divmod(c, 50)[1] == remain[1]):
                    # print(self.taskdata)
                    # 向网络请求数据分发任务图斑
                    res = requests.post(url=self.assign_url, json=self.taskdata, cookies=self.cookies, headers=self.headers, verify=False)
                    jdata = res.json()
                    print("[ {0} ]++第{1}波网络返回数据：{2},code:{3},error:{4},message:{5}！".format(self.hms_now(), flag, jdata['data'],jdata['code'],jdata['error'],jdata['message']))
                    return
                elif remain[0] >= flag and divmod(c, 50)[1] == 0:
                    # 向网络请求数据分发任务图斑
                    res = requests.post(url=self.assign_url, json=self.taskdata, cookies=self.cookies, headers=self.headers, verify=False)
                    jdata = res.json()
                    self.taskdata['tbbsms'] = []
                    print("[ {0} ]++第{1}波50个网络返回数据：{2},code:{3},error:{4},message:{5}！".format(self.hms_now(), flag, jdata['data'],jdata['code'],jdata['error'],jdata['message']))
                    flag += 1
                    continue
                c += 1
        else:
            print("[ %s ]**********匹配分发任务出现问题,请检查后重新运行！" % self.hms_now())
if __name__ == '__main__':
    tysd = Tysdgx()
    tysd.rtv_data()
    tysd.begin_assignment()
